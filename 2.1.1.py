import csv
from datetime import datetime
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

class SalaryDict:
    ''' Класс для для получения данных по зарплатам для выбранной вакансии
    '''
    def __init__(self):
        ''' Создает словари для дальнейшей работы

        salary_dict (dict) : словарь для всех зарплат в файле, где ключ - это название города, а значение - массив зарплат
        average_salary_dict (dict) : словарь для записи средних зарплат
        '''
        self.salary_dict = {}
        self.__average_salary_dict = {}

    def add_salary(self, key, salary):
        ''' Добавляет все поля salary в словарь по городам, если ключа нет в словаре, на место ключа записывает пустой массив

            key (str or int) : ключ для получения зарплаты в словаре
            salary (int or float) : зарплата, которую добавляет в словарь
        '''

        if self.salary_dict.get(key) is None:
            self.salary_dict[key] = []
        return self.salary_dict[key].append(salary)

    def get_average_salary(self):
        ''' Записывает в average_salary_dict среднюю зарплату, возвращает словарь, где ключ - это название города, а значение - средняя зп
        '''

        for key, value in self.salary_dict.items():
            self.__average_salary_dict[key] = int(mean(value))
        return self.__average_salary_dict

    def top_salary(self, big_cities):
        ''' Находит самые большие срдение зарплаты, записывает в словарь топ 10

            big_cities (array) : массив самых крупных городов
        '''

        self.get_average_salary()
        sorted_dict = dict(sorted(self.__average_salary_dict.items(), key=lambda x: x[1], reverse=True))
        big_salary_dict = {}
        for key, value in sorted_dict.items():
            if key in big_cities:
                big_salary_dict[key] = value
        return {x: big_salary_dict[x] for x in list(big_salary_dict)[:10]}

class CountDict:
    ''' Класс для получения количества вакансий для выбранной профессии'''

    def __init__(self):
        ''' length (int) : счетчик для получения количества вакансий
            count_dict (dict) : словарь, куда будут записываться данные по количеству
            big_cities (array) : массив самых крупных городов
            top_proportion_dict (dict) :
        '''
        self.length = 0
        self.count_dict = {}
        self.big_cities = []
        self.top_proportion_dict = {}

    def add(self, key):
        ''' Считает количество вакансий для городов и годов, записывает число в словарь

           key (str or int) : ключ для получения данных в словаре
        '''
        if self.count_dict.get(key) is None:
            self.count_dict[key] = 0
        self.count_dict[key] += 1
        self.length += 1
        return

    def get_proportion(self):
        ''' Выбирает самые крупные города
            Получает процент для каждого города, город с большим значением записывает в массив'''

        proportion_dict = {}
        for key, value in self.count_dict.items():
            proportion = value / self.length
            if proportion >= 0.01:
                self.big_cities.append(key)
                proportion_dict[key] = round(proportion, 4)
        sorted_dict = dict(sorted(proportion_dict.items(), key=lambda x: x[1], reverse=True))
        self.top_proportion_dict = {x: sorted_dict[x] for x in list(sorted_dict)[:10]}
        return

class Vacancy:
    ''' Класс для получения данных о вакансии'''
    def __init__(self, data):
        ''' Записывает в нужные поля все данные о вакансии, для зарплаты ищет среднее и переводит в рубли, время переводит в нужный формат

        data(array) : массив с одной вакансией
        '''

        if len(data) != 6:
            data = [data[0], data[6], data[7], data[9], data[10], data[11]]
        self.__dict_currency = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13,
                "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}
        self.job = data[0]
        self.salary = (float(data[1]) + float(data[2])) / 2 * self.__dict_currency[data[3]]
        self.city = data[4]
        self.year = int(datetime.strptime(data[5], "%Y-%m-%dT%H:%M:%S%z").strftime('%Y'))

class Report:
    ''' Класс для формирования таблицы'''
    def __init__(self):
        ''' Создает нужные перемнные, заголовки для таблицы, саму таблицу и страницы в excel'''

        self.name_columns_years = ['Год', 'Средняя зарплата', f'Средняя зарплата - {job}', 'Количество вакансий',
                              f'Количество вакансий - {job}']
        self.name_columns_cities = ['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий']
        self.columns = ['A', 'B', 'C', 'D', 'E']
        self.wb = Workbook()
        self.years = self.wb.active
        self.years.title = 'Статистика по годам'
        self.cities = self.wb.create_sheet('Статистика по городам')
        self.cities = self.wb['Статистика по городам']

        self.salary_year = SalaryDict()
        self.count_year = CountDict()
        self.job_salary_year = SalaryDict()
        self.job_count_year = CountDict()
        self.job_salary_city = SalaryDict()
        self.job_count_city = CountDict()

    def get_data(self, vacancies, job):
        ''' Добавляет все нужные данные в словари из вакансий

        vacancies(array): массив всех вакансий
        job(str) : нужная вакансия
        '''

        for vacancy in vacancies:
            self.salary_year.add_salary(vacancy.year, vacancy.salary)
            self.count_year.add(vacancy.year)
            self.job_salary_city.add_salary(vacancy.city, vacancy.salary)
            self.job_count_city.add(vacancy.city)
            if job in vacancy.job:
                self.job_salary_year.add_salary(vacancy.year, vacancy.salary)
                self.job_count_year.add(vacancy.year)
        if self.job_salary_year.salary_dict == {}:
            self.job_salary_year.salary_dict = {x: [0] for x in self.salary_year.salary_dict.keys()}
        if self.job_count_year.count_dict == {}:
            self.job_count_year.count_dict = {x: 0 for x in self.count_year.count_dict.keys()}
        self.job_count_city.get_proportion()
        return

    def set_border(self, ws, cell_range):
        ''' Рисует нужную границу для таблиц

        ws: страница excel книги
        cell_range(str) : диапозон для границ
        '''
        thin = Side(border_style="thin")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        return

    def write_column_names(self, list, names, interval1, interval2):
        ''' Записывает названия для колонок таблицы

        list: страница excel книги
        interval1, interval2(str) : диапозон для границ
        names(array): массив названий
        '''
        self.set_border(list, interval1)
        self.set_border(list, interval2)
        for i in range(1, len(names) + 1):
            list.cell(row=1, column=i).value = names[i - 1]
            list.cell(row=1, column=i).font = Font(bold = True)
        return

    def length(self, list):
        ''' Выставляет ширину колонок в зависимости от содержимого

        list: страница excel книги
        '''
        dims = {}
        for row in list.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            list.column_dimensions[col].width = value + 2
        return

    def get_dates_xl(self):
        ''' Записывает данные из словарей в ячейки таблицы, колонку с долей вакансий приводит к процентному формату'''

        salary_years = self.salary_year.get_average_salary()
        vac_years = self.count_year.count_dict
        prof_years = self.job_salary_year.get_average_salary()
        amount_years = self.job_count_year.count_dict
        salary_cities = self.job_salary_city.top_salary(self.job_count_city.big_cities)
        vac_cities = self.job_count_city.top_proportion_dict

        def write_row(dict, list, column1, column2):
            ''' Записывает данные из словарей по строкам со второй строки

            list: страница excel книги
            dict: словарь с нужными значениями
            column1(str): колонка таблицы для записи города или года
            column2(str): колонка таблицы для записи значений
            '''

            n = 2
            for i, k in dict.items():
                list[f"{column1}{n}"] = i
                list[f"{column2}{n}"] = k
                n += 1
            return

        for i in range(2,12):
            self.cities[f"E{i}"].number_format = FORMAT_PERCENTAGE_00

        write_row(salary_years, self.years, 'A', 'B')
        write_row(vac_years, self.years, 'A', 'D')
        write_row(prof_years, self.years, 'A', 'C')
        write_row(amount_years, self.years, 'A', 'E')
        write_row(salary_cities, self.cities, 'A', 'B')
        write_row(vac_cities, self.cities, 'D', 'E')
        return

    def print_result(self):
        ''' Выводит все словари с данными'''

        print(f"Динамика уровня зарплат по годам: {self.salary_year.get_average_salary()}")
        print(f"Динамика количества вакансий по годам: {self.count_year.count_dict}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {self.job_salary_year.get_average_salary()}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {self.job_count_year.count_dict}")
        print(
            f"Уровень зарплат по городам (в порядке убывания): {self.job_salary_city.top_salary(self.job_count_city.big_cities)}")
        print(f"Доля вакансий по городам (в порядке убывания): {self.job_count_city.top_proportion_dict}")
        return

    def write(self):
        ''' Записывает данные в таблицу, форматирует ее по примеру, сохраняет ее'''

        self.date = self.get_dates_xl
        self.write_column_names(self.years, self.name_columns_years, 'A1:E17', 'A1:E17')
        self.write_column_names(self.cities, self.name_columns_cities, 'A1:B11', 'D1:E11')

        self.get_dates_xl()

        self.length(self.years)
        self.length(self.cities)

        self.wb.save('report.xlsx')
        return

def csv_reader(file_name):
    ''' Получает данные о файле и считывает его, проверяет, чтобы файл не был пустым, возвращает лист с вакансиями'''
    file = open(file_name, encoding = 'utf_8_sig')
    reader = csv.reader(file)
    list_data = list(filter(lambda x: '' not in x, reader))
    if len(list_data) == 0:
        return print("Пустой файл")
    if len(list_data[1:]) == 0:
        return print('Het данных')
    return list_data[1:]

file_name = input("Введите название файла: ")
job = input("Введите название профессии: ")
file = open(file_name, encoding="utf_8_sig")
data = csv_reader(file_name)

if data is not None:
    '''Проверяет чтобы вакансии не были пустыми, вызывает все методы, записывает в таблицу данные, выводит результат'''
    vacancies = [Vacancy(x) for x in data]
    result = Report()
    result.get_data(vacancies, job)
    result.write()
    result.print_result()
